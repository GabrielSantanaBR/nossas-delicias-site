from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from store.financial_models import ProductCostProfile
from store.models import PriceTable, ProductPrice


def norm(value):
    text = unicodedata.normalize('NFD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(text.strip().lower().split())


def decimal_or_none(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value).replace('R$', '').replace(' ', '').replace(',', '.'))
    except (InvalidOperation, ValueError):
        return None


def sale_unit(value):
    value = norm(value)
    if 'fatia' in value:
        return 'slice'
    if 'grama' in value or 'porcao' in value:
        return 'portion'
    if 'caixa' in value or 'kit' in value:
        return 'box'
    if 'unidade' in value:
        return 'unit'
    return 'other'


class Command(BaseCommand):
    help = 'Sincroniza custos e preços da aba PRECIFICAÇÃO da planilha Nossas Delícias.'

    def add_arguments(self, parser):
        parser.add_argument('workbook', help='Caminho do arquivo .xlsx')
        parser.add_argument('--sheet', default='PRECIFICAÇÃO')
        parser.add_argument('--dry-run', action='store_true', help='Mostra o que seria alterado sem gravar.')

    def handle(self, *args, **options):
        path = Path(options['workbook'])
        if not path.exists() or path.suffix.lower() != '.xlsx':
            raise CommandError('Informe um arquivo .xlsx existente.')

        wb = load_workbook(path, data_only=True, read_only=True)
        if options['sheet'] not in wb.sheetnames:
            raise CommandError(f'Aba {options["sheet"]!r} não encontrada. Disponíveis: {", ".join(wb.sheetnames)}')
        ws = wb[options['sheet']]

        header_row = None
        headers = {}
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = [norm(v) for v in row]
            if 'codigo' in normalized and 'produto' in normalized and any('custo unitario' == v for v in normalized):
                header_row = idx
                headers = {norm(value): pos for pos, value in enumerate(row) if value not in (None, '')}
                break
        if header_row is None:
            raise CommandError('Não encontrei o cabeçalho da PRECIFICAÇÃO (Código/Produto/Custo unitário).')

        required = ['codigo', 'produto', 'rendimento (qtd.)', 'custo total', 'custo unitario', 'preco cafeteria', 'preco cliente', 'venda ativa?']
        missing = [name for name in required if name not in headers]
        if missing:
            raise CommandError(f'Colunas obrigatórias ausentes: {", ".join(missing)}')

        cafe_table, _ = PriceTable.objects.get_or_create(name='Cafeterias', kind='cafe', defaults={'active': True})
        retail_table, _ = PriceTable.objects.get_or_create(name='Clientes', kind='retail', defaults={'active': True})

        stats = {'updated': 0, 'unmapped': 0, 'prices': 0, 'skipped': 0}
        unmapped = []

        @transaction.atomic
        def apply_rows():
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                code = str(row[headers['codigo']] or '').strip().upper()
                if not code:
                    continue
                product_name = str(row[headers['produto']] or '').strip()
                active_text = norm(row[headers['venda ativa?']])
                active_sale = active_text in {'sim', 's', 'yes', 'true', '1'}
                profile = ProductCostProfile.objects.select_related('product').filter(sku__iexact=code).first()
                if not profile:
                    stats['unmapped'] += 1
                    unmapped.append(f'{code} — {product_name}')
                    continue

                yield_qty = decimal_or_none(row[headers['rendimento (qtd.)']])
                production_cost = decimal_or_none(row[headers['custo total']])
                sheet_unit_cost = decimal_or_none(row[headers['custo unitario']])
                if not yield_qty or yield_qty <= 0 or production_cost is None:
                    stats['skipped'] += 1
                    continue

                profile.sale_unit = sale_unit(row[headers.get('unidade de venda')]) if 'unidade de venda' in headers else profile.sale_unit
                profile.yield_quantity = yield_qty
                profile.production_cost = production_cost
                profile.active = active_sale
                profile.source_reference = f'{path.name} / {code}'[:120]
                if not options['dry_run']:
                    profile.save()
                # Profile.save recalculates unit_cost. Compare to the workbook only as a warning.
                calculated = (production_cost / yield_qty).quantize(Decimal('0.0001'))
                if sheet_unit_cost is not None and abs(calculated - sheet_unit_cost) > Decimal('0.01'):
                    self.stderr.write(self.style.WARNING(f'{code}: custo unitário recalculado {calculated} difere da planilha {sheet_unit_cost}'))
                stats['updated'] += 1

                for table, column in ((cafe_table, 'preco cafeteria'), (retail_table, 'preco cliente')):
                    value = decimal_or_none(row[headers[column]])
                    if value is None or value <= 0:
                        continue
                    stats['prices'] += 1
                    if not options['dry_run']:
                        ProductPrice.objects.update_or_create(
                            product=profile.product,
                            table=table,
                            min_quantity=1,
                            defaults={'unit_price': value},
                        )

            if options['dry_run']:
                transaction.set_rollback(True)

        apply_rows()
        mode = 'SIMULAÇÃO' if options['dry_run'] else 'IMPORTAÇÃO'
        self.stdout.write(self.style.SUCCESS(
            f'{mode}: {stats["updated"]} custos mapeados, {stats["prices"]} preços lidos, '
            f'{stats["unmapped"]} códigos sem produto mapeado, {stats["skipped"]} linhas incompletas.'
        ))
        if unmapped:
            self.stdout.write('\nMapeie estes códigos em “Custos dos produtos” antes da próxima importação:')
            for item in unmapped[:50]:
                self.stdout.write(f'  - {item}')
            if len(unmapped) > 50:
                self.stdout.write(f'  ... e mais {len(unmapped)-50}.')
