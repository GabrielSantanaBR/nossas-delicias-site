import hashlib
import io
import unicodedata
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .financial_models import BusinessExpense, ProductCostProfile
from .financial_services import sales_report
from .management_models import (
    FinancialSettings,
    FixedCost,
    Ingredient,
    IngredientPriceHistory,
    InventoryMovement,
    Recipe,
    SpreadsheetImportBatch,
)
from .management_services import cashflow_summary, pricing_health, sync_recipe_product_cost
from .models import CafeAccount, PriceTable, ProductPrice

HEADER_FILL = '3A211A'
MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 5000


def _norm(value):
    text = str(value or '').strip().upper()
    text = ''.join(ch for ch in unicodedata.normalize('NFKD', text) if not unicodedata.combining(ch))
    return ' '.join(text.split())


def _decimal(value):
    if value is None or value == '':
        return None
    try:
        if isinstance(value, str):
            text = value.replace('\xa0', ' ').replace('R$', '').replace('%', '').strip().replace(' ', '')
            if ',' in text:
                text = text.replace('.', '').replace(',', '.')
            value = text
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _yes(value):
    return _norm(value) in {'SIM', 'S', 'YES', 'ATIVO', 'TRUE', '1'}


def _ingredient_unit(value):
    normalized = _norm(value)
    return {
        'G': 'g', 'GRAMA': 'g', 'GRAMAS': 'g',
        'ML': 'ml', 'MILILITRO': 'ml', 'MILILITROS': 'ml',
        'UN': 'un', 'UNIDADE': 'un', 'UNIDADES': 'un',
        'KG': 'kg', 'QUILOGRAMA': 'kg', 'QUILOGRAMAS': 'kg',
        'L': 'l', 'LITRO': 'l', 'LITROS': 'l',
    }.get(normalized, 'other')


def _sale_unit(value):
    normalized = _norm(value)
    return {
        'UN': 'unit', 'UNIDADE': 'unit', 'UNIDADES': 'unit',
        'FATIA': 'slice', 'FATIAS': 'slice',
        'GRAMA/PORCAO': 'portion', 'GRAMA / PORCAO': 'portion', 'PORCAO': 'portion',
        'CAIXA/KIT': 'box', 'CAIXA / KIT': 'box', 'CAIXA': 'box', 'KIT': 'box',
    }.get(normalized, 'other')


def _date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for pattern in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(value.strip(), pattern).date()
            except ValueError:
                pass
    return None


def _find_header(ws, required, max_rows=80):
    required = {_norm(item) for item in required}
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row), values_only=True), start=1):
        normalized = {_norm(cell) for cell in row if cell not in (None, '')}
        if required.issubset(normalized):
            return idx
    return None


def _header_map(ws, row_number):
    values = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
    return {_norm(value): index for index, value in enumerate(values) if value not in (None, '')}


def _cell(row, mapping, name):
    index = mapping.get(_norm(name))
    return row[index] if index is not None and index < len(row) else None


def _sheet(wb, name):
    target = _norm(name)
    for ws in wb.worksheets:
        if _norm(ws.title) == target:
            return ws
    return None


def _read_bytes(uploaded_file):
    uploaded_file.seek(0)
    data = uploaded_file.read(MAX_XLSX_BYTES + 1)
    uploaded_file.seek(0)
    if len(data) > MAX_XLSX_BYTES:
        raise ValueError('A planilha ultrapassa o limite de 10 MB.')
    return data


def _validate_xlsx_archive(raw):
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise ValueError('A planilha contém arquivos internos demais.')
            names = {entry.filename for entry in entries}
            if '[Content_Types].xml' not in names or 'xl/workbook.xml' not in names:
                raise ValueError('O arquivo não possui uma estrutura .xlsx válida.')
            total = 0
            for entry in entries:
                name = entry.filename.replace('\\', '/')
                if name.startswith('/') or '..' in name.split('/'):
                    raise ValueError('A planilha contém caminhos internos inválidos.')
                total += entry.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise ValueError('A planilha expandida é grande demais para processamento seguro.')
                if entry.file_size > 2 * 1024 * 1024 and entry.compress_size and entry.file_size / entry.compress_size > 500:
                    raise ValueError('A planilha possui compactação interna suspeita.')
    except zipfile.BadZipFile as exc:
        raise ValueError('O arquivo enviado não é uma planilha .xlsx válida.') from exc


@transaction.atomic
def import_management_workbook(uploaded_file, user=None):
    raw = _read_bytes(uploaded_file)
    _validate_xlsx_archive(raw)
    checksum = hashlib.sha256(raw).hexdigest()
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True, keep_links=False)
    summary = {
        'ingredients_created': 0, 'ingredients_updated': 0,
        'recipes_created': 0, 'recipes_updated': 0,
        'prices_updated': 0, 'skipped_rows': 0,
        'warnings': [],
    }

    base = _sheet(wb, 'BASE DE PREÇOS')
    if base:
        header_row = _find_header(base, ['Código', 'Ingrediente padrão', 'Preço da embalagem', 'Qtd. da embalagem'])
        if header_row:
            headers = _header_map(base, header_row)
            for row in base.iter_rows(min_row=header_row + 1, values_only=True):
                code = str(_cell(row, headers, 'Código') or '').strip()
                name = str(_cell(row, headers, 'Ingrediente padrão') or '').strip()
                if not code.startswith('ING-') or not name:
                    continue
                price = _decimal(_cell(row, headers, 'Preço da embalagem')) or Decimal('0')
                quantity = _decimal(_cell(row, headers, 'Qtd. da embalagem')) or Decimal('1')
                if price < 0 or quantity <= 0:
                    summary['skipped_rows'] += 1
                    continue
                last_update = _date_value(_cell(row, headers, 'Última atualização')) or date.today()
                defaults = {
                    'name': name[:140],
                    'category': str(_cell(row, headers, 'Categoria') or '')[:100],
                    'package_price': price,
                    'package_quantity': quantity,
                    'base_unit': _ingredient_unit(_cell(row, headers, 'Unidade base')),
                    'supplier': str(_cell(row, headers, 'Fornecedor') or '')[:140],
                    'notes': str(_cell(row, headers, 'Observações') or ''),
                    'aliases': str(_cell(row, headers, 'Nomes encontrados') or ''),
                    'active': _norm(_cell(row, headers, 'Status')) != 'INATIVO',
                    'last_price_update': last_update,
                }
                ingredient = Ingredient.objects.filter(code=code).first()
                old_price = ingredient.package_price if ingredient else None
                old_quantity = ingredient.package_quantity if ingredient else None
                ingredient, created = Ingredient.objects.update_or_create(code=code, defaults=defaults)
                summary['ingredients_created' if created else 'ingredients_updated'] += 1
                if created or old_price != ingredient.package_price or old_quantity != ingredient.package_quantity:
                    IngredientPriceHistory.objects.create(
                        ingredient=ingredient,
                        package_price=ingredient.package_price,
                        package_quantity=ingredient.package_quantity,
                        unit_cost=ingredient.unit_cost,
                        supplier=ingredient.supplier,
                        source=getattr(uploaded_file, 'name', 'Planilha importada')[:160],
                        effective_date=ingredient.last_price_update,
                    )
        else:
            summary['warnings'].append('Cabeçalho da BASE DE PREÇOS não foi reconhecido.')
    else:
        summary['warnings'].append('Aba BASE DE PREÇOS não encontrada.')

    pricing = _sheet(wb, 'PRECIFICAÇÃO')
    if pricing:
        for row in pricing.iter_rows(min_row=1, max_row=min(pricing.max_row, 15), values_only=True):
            for index, value in enumerate(row):
                if _norm(value) == 'MARGEM DESEJADA':
                    candidate = next((_decimal(v) for v in row[index + 1:] if _decimal(v) is not None), None)
                    if candidate is not None:
                        if candidate <= 1:
                            candidate *= 100
                        settings = FinancialSettings.current()
                        settings.desired_margin_percent = max(Decimal('0'), min(candidate, Decimal('95')))
                        settings.save(update_fields=['desired_margin_percent', 'updated_at'])
                    break

        header_row = _find_header(pricing, ['Código', 'Categoria', 'Produto', 'Rendimento (qtd.)', 'Custo total', 'Custo unitário'])
        if header_row:
            headers = _header_map(pricing, header_row)
            retail_table, _ = PriceTable.objects.get_or_create(name='Cliente padrão', kind='retail', defaults={'active': True})
            cafe_table, _ = PriceTable.objects.get_or_create(name='Cafeteria padrão', kind='cafe', defaults={'active': True})
            for row in pricing.iter_rows(min_row=header_row + 1, values_only=True):
                code = str(_cell(row, headers, 'Código') or '').strip()
                name = str(_cell(row, headers, 'Produto') or '').strip()
                if not code.startswith('REC-') or not name:
                    continue
                yield_quantity = _decimal(_cell(row, headers, 'Rendimento (qtd.)')) or Decimal('1')
                production_cost = _decimal(_cell(row, headers, 'Custo total')) or Decimal('0')
                if yield_quantity <= 0 or production_cost < 0:
                    summary['skipped_rows'] += 1
                    continue
                profile = ProductCostProfile.objects.filter(sku=code).select_related('product').first()
                recipe = Recipe.objects.filter(code=code).first()
                product = profile.product if profile else (recipe.product if recipe else None)
                defaults = {
                    'name': name[:180],
                    'category': str(_cell(row, headers, 'Categoria') or '')[:100],
                    'sale_unit': _sale_unit(_cell(row, headers, 'Unidade de venda')),
                    'yield_quantity': yield_quantity,
                    'imported_production_cost': production_cost,
                    'product': product,
                    'active': _yes(_cell(row, headers, 'Venda ativa?')),
                    'source_reference': f'{getattr(uploaded_file, "name", "Planilha")} / {code}'[:160],
                }
                recipe, created = Recipe.objects.update_or_create(code=code, defaults=defaults)
                summary['recipes_created' if created else 'recipes_updated'] += 1
                if product:
                    sync_recipe_product_cost(recipe)
                    cafe_price = _decimal(_cell(row, headers, 'Preço cafeteria'))
                    client_price = _decimal(_cell(row, headers, 'Preço cliente'))
                    if cafe_price and cafe_price > 0:
                        ProductPrice.objects.update_or_create(product=product, table=cafe_table, min_quantity=product.min_quantity, defaults={'unit_price': cafe_price})
                        summary['prices_updated'] += 1
                    if client_price and client_price > 0:
                        ProductPrice.objects.update_or_create(product=product, table=retail_table, min_quantity=product.min_quantity, defaults={'unit_price': client_price})
                        summary['prices_updated'] += 1
        else:
            summary['warnings'].append('Cabeçalho da PRECIFICAÇÃO não foi reconhecido.')
    else:
        summary['warnings'].append('Aba PRECIFICAÇÃO não encontrada.')

    status = 'partial' if summary['warnings'] else 'success'
    batch = SpreadsheetImportBatch.objects.create(
        filename=getattr(uploaded_file, 'name', 'planilha.xlsx')[:180],
        sha256=checksum,
        source_version='Planilha Nossas Delícias',
        status=status,
        imported_by=user if getattr(user, 'is_authenticated', False) else None,
        ingredients_created=summary['ingredients_created'],
        ingredients_updated=summary['ingredients_updated'],
        recipes_created=summary['recipes_created'],
        recipes_updated=summary['recipes_updated'],
        prices_updated=summary['prices_updated'],
        skipped_rows=summary['skipped_rows'],
        summary={'warnings': summary['warnings']},
    )
    summary['batch'] = batch
    return summary


def _style_sheet(ws, freeze='A2'):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 26
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(vertical='center')
    money_headers = {'PREÇO DA EMBALAGEM', 'CUSTO POR UNIDADE', 'PREÇO CAFETERIA', 'PREÇO CLIENTE', 'LUCRO/UN. CAFÉ', 'LUCRO/UN. CLIENTE', 'PREÇO RECOMENDADO', 'PREÇO UNITÁRIO', 'FATURAMENTO', 'CUSTO UNITÁRIO', 'CUSTO TOTAL', 'LUCRO', 'VALOR', 'VALOR MENSAL'}
    date_headers = {'DATA', 'ÚLTIMA ATUALIZAÇÃO', 'INÍCIO', 'FIM'}
    percent_headers = {'MARGEM', 'MARGEM CAFÉ', 'MARGEM CLIENTE'}
    headers = {str(cell.value or '').strip().upper(): cell.column for cell in ws[1]}
    for label, column in headers.items():
        for target in (ws.cell(row=row, column=column) for row in range(2, ws.max_row + 1)):
            if label in money_headers:
                target.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
            elif label in date_headers:
                target.number_format = 'dd/mm/yyyy'
            elif label in percent_headers:
                target.number_format = '0.00"%"'
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(max((len(str(cell.value or '')) for cell in column), default=8) + 2, 34)
        ws.column_dimensions[letter].width = max(width, 11)


def _append_money(value):
    return float(Decimal(value or 0).quantize(Decimal('0.01')))


def _safe_sheet_title(value, used):
    cleaned = ''.join('-' if char in '[]:*?/\\' else char for char in str(value or '')).strip() or 'SEM NOME'
    base = cleaned[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        marker = f' {suffix}'
        candidate = f'{base[:31 - len(marker)]}{marker}'
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def build_management_workbook(start, end):
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_titles = set()

    health = pricing_health()
    sales = sales_report(start, end)
    cash = cashflow_summary(start, end)

    ws = wb.create_sheet(_safe_sheet_title('PAINEL', used_sheet_titles))
    ws.append(['PAINEL DE GESTÃO', 'Valor'])
    ws.append(['Período', f'{start:%d/%m/%Y} a {end:%d/%m/%Y}'])
    ws.append(['Receitas ativas', health['active_recipes']])
    ws.append(['Preços cafeteria preenchidos', health['cafe_prices']])
    ws.append(['Preços cliente preenchidos', health['client_prices']])
    ws.append(['Margem desejada (%)', float(health['desired_margin'])])
    ws.append(['Itens vendidos', sales['totals']['items']])
    ws.append(['Faturamento de itens', _append_money(sales['totals']['revenue'])])
    ws.append(['Recebido', _append_money(cash['cash_in'])])
    ws.append(['A receber', _append_money(cash['receivables']['total'])])
    ws.append(['Lucro bruto', _append_money(sales['totals']['profit'])])
    ws.append(['Despesas pagas', _append_money(cash['cash_out'])])
    ws.append(['Caixa líquido', _append_money(cash['net_cash'])])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('BASE DE PREÇOS', used_sheet_titles))
    ws.append(['Código', 'Ingrediente padrão', 'Categoria', 'Preço da embalagem', 'Qtd. da embalagem', 'Unidade base', 'Custo por unidade', 'Estoque atual', 'Estoque mínimo', 'Status', 'Última atualização', 'Fornecedor', 'Observações', 'Nomes encontrados'])
    for item in Ingredient.objects.all():
        ws.append([item.code, item.name, item.category, _append_money(item.package_price), float(item.package_quantity), item.base_unit, float(item.unit_cost), float(item.stock_balance), float(item.minimum_stock), 'ATIVO' if item.active else 'INATIVO', item.last_price_update, item.supplier, item.notes, item.aliases])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('PRECIFICAÇÃO', used_sheet_titles))
    ws.append(['Código', 'Categoria', 'Produto', 'Unidade de venda', 'Rendimento (qtd.)', 'Custo total', 'Custo unitário', 'Preço cafeteria', 'Preço cliente', 'Lucro/un. café', 'Margem café', 'Lucro/un. cliente', 'Margem cliente', 'Preço recomendado', 'Situação', 'Venda ativa?'])
    for row in health['rows']:
        recipe = row['recipe']
        ws.append([recipe.code, recipe.category, recipe.name, recipe.get_sale_unit_display(), float(recipe.yield_quantity), float(row['production_cost']), float(row['unit_cost']), None if row['cafe_price'] is None else float(row['cafe_price']), None if row['client_price'] is None else float(row['client_price']), None if row['cafe_profit'] is None else float(row['cafe_profit']), None if row['cafe_margin'] is None else float(row['cafe_margin']), None if row['client_profit'] is None else float(row['client_profit']), None if row['client_margin'] is None else float(row['client_margin']), None if row['recommended_price'] is None else float(row['recommended_price']), row['status'], 'SIM' if recipe.active else 'NÃO'])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('RECEITAS', used_sheet_titles))
    ws.append(['Código', 'Receita', 'Categoria', 'Ativa', 'Rendimento', 'Unidade de venda', 'Código ingrediente', 'Ingrediente / insumo', 'Quantidade usada', 'Unidade base', 'Custo do ingrediente', 'Custo da linha', 'Custo total da receita', 'Custo por unidade', 'Produto vinculado'])
    recipes = Recipe.objects.select_related('product').prefetch_related('ingredients__ingredient').all()
    for recipe in recipes:
        lines = list(recipe.ingredients.all()) or [None]
        for line in lines:
            ingredient = line.ingredient if line else None
            line_cost = line.total_cost if line else None
            ws.append([
                recipe.code, recipe.name, recipe.category, 'SIM' if recipe.active else 'NÃO', float(recipe.yield_quantity), recipe.get_sale_unit_display(),
                ingredient.code if ingredient else '', ingredient.name if ingredient else 'Composição pendente', float(line.quantity_used) if line else None,
                ingredient.base_unit if ingredient else '', float(ingredient.unit_cost) if ingredient else None, float(line_cost) if line_cost is not None else None,
                float(recipe.production_cost), float(recipe.unit_cost), recipe.product.name if recipe.product else '',
            ])
    _style_sheet(ws)

    def sales_sheet(title, order_type=None, cafe=None):
        report = sales_report(start, end, order_type=order_type, cafe=cafe)
        sheet = wb.create_sheet(_safe_sheet_title(title, used_sheet_titles))
        sheet.append(['Data', 'Pedido/nota', 'Canal', 'Cliente/cafeteria', 'Código', 'Produto', 'Quantidade', 'Preço unitário', 'Faturamento', 'Custo unitário', 'Custo total', 'Lucro', 'Margem', 'Pagamento', 'Mês', 'Ano', 'Observações'])
        for snap in report['rows']:
            order = snap.order_item.order
            cafe = getattr(order.customer, 'cafe_account', None)
            note = getattr(order, 'cafe_delivery_note', None)
            d = order.delivery_date or order.created_at.date()
            sheet.append([d, note.note_number if note else str(order.public_id)[:8].upper(), order.order_type, cafe.business_name if cafe else (order.customer.get_full_name() or order.customer.username), snap.sku, snap.product_name, snap.quantity, float(snap.unit_price), float(snap.revenue), None if snap.unit_cost is None else float(snap.unit_cost), None if snap.total_cost is None else float(snap.total_cost), None if snap.profit is None else float(snap.profit), None if snap.margin_percent is None else float(snap.margin_percent), 'Pago' if order.payments.filter(status='approved').exists() else 'Pendente', d.month, d.year, order.customer_note])
        _style_sheet(sheet)

    sales_sheet('VENDAS CLIENTES', 'retail')
    sales_sheet('VENDAS CAFETERIAS', 'cafe')
    for cafe in CafeAccount.objects.select_related('user').order_by('business_name', 'pk'):
        sales_sheet(f'CAFÉ - {cafe.business_name}', 'cafe', cafe=cafe)
    sales_sheet('VENDAS EVENTOS', 'event')

    ws = wb.create_sheet(_safe_sheet_title('ANÁLISE DE VENDAS', used_sheet_titles))
    ws.append(['Cafeteria/cliente', 'Itens', 'Pedidos', 'Faturamento', 'Custo', 'Lucro', 'Margem'])
    for row in sales['by_cafe']:
        ws.append([row['name'], row['quantity'], row.get('order_count', 0), _append_money(row['revenue']), _append_money(row['cost']), _append_money(row['profit']), float(row['margin_percent'])])
    ws.append([])
    ws.append(['Mês', 'Itens', 'Faturamento', 'Custo', 'Lucro', 'Margem'])
    for row in sales['by_month']:
        ws.append([row['name'], row['quantity'], _append_money(row['revenue']), _append_money(row['cost']), _append_money(row['profit']), float(row['margin_percent'])])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('DESPESAS', used_sheet_titles))
    ws.append(['Data', 'Categoria', 'Descrição', 'Fornecedor', 'Valor', 'Status', 'Pagamento', 'Observações'])
    for expense in BusinessExpense.objects.filter(date__range=(start, end)).order_by('date'):
        ws.append([expense.date, expense.get_category_display(), expense.description, expense.supplier, _append_money(expense.amount), expense.get_payment_status_display(), expense.payment_method, expense.notes])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('CUSTOS FIXOS', used_sheet_titles))
    ws.append(['Nome', 'Categoria', 'Valor mensal', 'Dia vencimento', 'Ativo', 'Início', 'Fim', 'Observações'])
    for cost in FixedCost.objects.all():
        ws.append([cost.name, cost.get_category_display(), _append_money(cost.monthly_amount), cost.due_day, 'SIM' if cost.active else 'NÃO', cost.start_date, cost.end_date, cost.notes])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('ESTOQUE', used_sheet_titles))
    ws.append(['Data', 'Código', 'Ingrediente', 'Tipo', 'Quantidade', 'Custo unitário', 'Referência', 'Observações'])
    for movement in InventoryMovement.objects.select_related('ingredient').filter(date__range=(start, end)).order_by('date', 'ingredient__name'):
        ws.append([movement.date, movement.ingredient.code, movement.ingredient.name, movement.get_movement_type_display(), float(movement.quantity_delta), float(movement.unit_cost_snapshot or 0), movement.reference, movement.notes])
    _style_sheet(ws)

    ws = wb.create_sheet(_safe_sheet_title('FLUXO DE CAIXA', used_sheet_titles))
    ws.append(['Indicador', 'Valor'])
    ws.append(['Entradas recebidas', _append_money(cash['cash_in'])])
    ws.append(['Saídas pagas', _append_money(cash['cash_out'])])
    ws.append(['Saldo líquido', _append_money(cash['net_cash'])])
    ws.append(['Contas a receber', _append_money(cash['receivables']['total'])])
    ws.append(['A receber vencido', _append_money(cash['receivables']['overdue'])])
    ws.append(['Contas a pagar', _append_money(cash['payables']['total'])])
    ws.append(['Custo fixo previsto', _append_money(cash['fixed_cost_forecast']['total'])])
    _style_sheet(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
