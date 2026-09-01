from decimal import Decimal

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from .financial_models import ProductCostProfile
from .financial_services import refresh_order_financials
from .management_models import FinancialSettings, FixedCost, Ingredient, InventoryMovement, Recipe, RecipeIngredient
from .management_services import simulate_price, sync_recipe_product_cost
from .models import CafeAccount, Category, Order, Product
from .spreadsheet_io import build_management_workbook


class ManagementCalculationTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Brownies', slug='brownies')
        self.product = Product.objects.create(
            category=self.category,
            name='Brownie Tradicional',
            slug='brownie-tradicional',
            description='Teste',
            image='products/brownie.jpg',
        )

    def test_ingredient_unit_cost_and_stock_movement(self):
        ingredient = Ingredient.objects.create(
            code='ING-001', name='Chocolate 100%', package_price=Decimal('30.00'),
            package_quantity=Decimal('500'), base_unit='g', minimum_stock=Decimal('100'),
        )
        self.assertEqual(ingredient.unit_cost, Decimal('0.060000'))
        InventoryMovement.objects.create(ingredient=ingredient, movement_type='purchase', quantity_delta=Decimal('1000'))
        InventoryMovement.objects.create(ingredient=ingredient, movement_type='production', quantity_delta=Decimal('250'))
        self.assertEqual(ingredient.stock_balance, Decimal('750.0000'))

    def test_recipe_cost_uses_current_ingredient_base_and_syncs_product_profile(self):
        ingredient = Ingredient.objects.create(
            code='ING-001', name='Chocolate 100%', package_price=Decimal('30.00'),
            package_quantity=Decimal('500'), base_unit='g',
        )
        recipe = Recipe.objects.create(
            code='REC-001', name='Brownie Tradicional', category='Brownies', sale_unit='unit',
            yield_quantity=Decimal('10'), extra_cost=Decimal('5.00'), product=self.product,
        )
        RecipeIngredient.objects.create(recipe=recipe, ingredient=ingredient, quantity_used=Decimal('500'))
        self.assertEqual(recipe.production_cost, Decimal('35.0000'))
        self.assertEqual(recipe.unit_cost, Decimal('3.5000'))
        sync_recipe_product_cost(recipe)
        profile = ProductCostProfile.objects.get(product=self.product)
        self.assertEqual(profile.unit_cost, Decimal('3.5000'))

    def test_price_simulator_matches_margin_and_break_even_rules(self):
        settings = FinancialSettings.current()
        settings.desired_margin_percent = Decimal('30')
        settings.payment_fee_percent = Decimal('0')
        settings.tax_percent = Decimal('0')
        settings.contingency_percent = Decimal('0')
        settings.save()
        FixedCost.objects.create(name='Estrutura', monthly_amount=Decimal('290.00'), due_day=1)
        recipe = Recipe.objects.create(
            code='REC-002', name='Teste', category='Brownies', sale_unit='unit',
            yield_quantity=Decimal('10'), imported_production_cost=Decimal('70.00'),
        )
        result = simulate_price(recipe, current_price=Decimal('9.00'), desired_margin=Decimal('30'), increase_percent=Decimal('10'), quantity=100)
        self.assertEqual(result['recommended_price'], Decimal('10.00'))
        self.assertEqual(result['new_price'], Decimal('9.90'))
        self.assertEqual(result['extra_total'], Decimal('90.00'))
        self.assertEqual(result['break_even_units'], 100)
        self.assertEqual(result['break_even_revenue'], Decimal('990.00'))


class SpreadsheetRoundTripTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('gestor', 'gestor@example.com', 'StrongPassword-123!', is_staff=True)
        category = Category.objects.create(name='Brownies', slug='brownies')
        self.product = Product.objects.create(
            category=category, name='Brownie Tradicional', slug='brownie-tradicional',
            description='Teste', image='products/brownie.jpg',
        )
        ProductCostProfile.objects.create(
            product=self.product, sku='REC-001', sale_unit='unit', yield_quantity=10,
            production_cost=Decimal('40.00'), source_reference='Teste',
        )

    def test_export_contains_management_sheets(self):
        Ingredient.objects.create(code='ING-001', name='Chocolate', package_price=30, package_quantity=500, base_unit='g')
        stream = build_management_workbook(timezone.localdate().replace(day=1), timezone.localdate())
        workbook = load_workbook(stream, read_only=True, data_only=True)
        expected = {'PAINEL', 'BASE DE PREÇOS', 'PRECIFICAÇÃO', 'RECEITAS', 'VENDAS CLIENTES', 'VENDAS CAFETERIAS', 'VENDAS EVENTOS', 'ANÁLISE DE VENDAS', 'DESPESAS', 'CUSTOS FIXOS', 'ESTOQUE', 'FLUXO DE CAIXA'}
        self.assertTrue(expected.issubset(set(workbook.sheetnames)))

    def test_export_creates_one_safe_sheet_for_every_cafe(self):
        first_user = User.objects.create_user('cafe-1', 'cafe1@example.com', 'StrongPassword-456!')
        second_user = User.objects.create_user('cafe-2', 'cafe2@example.com', 'StrongPassword-789!')
        first = CafeAccount.objects.create(user=first_user, business_name='Café / Zona Sul: Copacabana')
        CafeAccount.objects.create(user=second_user, business_name='Café / Zona Sul: Copacabana')
        order = Order.objects.create(
            customer=first.user, order_type='cafe', status='paid', delivery_date=timezone.localdate(),
            subtotal=Decimal('20.00'), total=Decimal('20.00'),
        )
        order.items.create(product=self.product, quantity=2, unit_price=Decimal('10.00'))
        refresh_order_financials(order)
        stream = build_management_workbook(timezone.localdate().replace(day=1), timezone.localdate())
        workbook = load_workbook(stream, read_only=True, data_only=True)
        cafe_sheets = [name for name in workbook.sheetnames if name.startswith('CAFÉ - ')]
        self.assertEqual(len(cafe_sheets), 2)
        self.assertEqual(len(set(name.casefold() for name in cafe_sheets)), 2)
        self.assertTrue(all(len(name) <= 31 and '/' not in name and ':' not in name for name in cafe_sheets))
        self.assertEqual(sum(workbook[name].max_row - 1 for name in cafe_sheets), 1)

    def test_management_center_requires_verified_admin(self):
        self.client.force_login(self.user)
        response = self.client.get('/gestao/')
        self.assertEqual(response.status_code, 403)

    def test_health_endpoint_checks_dependencies(self):
        response = self.client.get('/health/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')


class NativeCatalogTests(TestCase):
    def test_native_catalog_is_complete_and_idempotent(self):
        call_command('load_native_catalog', verbosity=0)
        self.assertEqual(Ingredient.objects.count(), 80)
        self.assertEqual(Recipe.objects.count(), 64)
        self.assertEqual(RecipeIngredient.objects.count(), 395)
        self.assertGreater(Recipe.objects.get(code='REC-001').production_cost, Decimal('0'))
        incomplete = Recipe.objects.get(code='REC-021')
        self.assertTrue(incomplete.active)
        self.assertFalse(incomplete.ingredients.exists())
        call_command('load_native_catalog', verbosity=0)
        self.assertEqual((Ingredient.objects.count(), Recipe.objects.count(), RecipeIngredient.objects.count()), (80, 64, 395))
